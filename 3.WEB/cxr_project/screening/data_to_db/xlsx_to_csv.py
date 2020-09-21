# 엑셀 파일을 csv파일로 변환
import openpyxl
import csv

xlsx_path = r"C:\Users\smile\ai_team\폐브리즈\CXR_AI\3.WEB\cxr_project\data_to_db\선별진료소.xlsx"
wb = openpyxl.load_workbook(xlsx_path)
sheet = wb['선별진료소(전체)']
max_row = sheet.max_row

f = open('screening_center.csv', 'w', encoding='utf-8', newline='')
wr = csv.writer(f)

for i in range(2, max_row+1):
    city = sheet.cell(i, 3).value
    town = sheet.cell(i, 4).value
    institutions = sheet.cell(i, 5).value
    address = sheet.cell(i, 6).value
    phone = sheet.cell(i, 10).value
    specimen = 'True' if sheet.cell(i, 2).value == '○' else 'False'
    
    wr.writerow([city, town, institutions, address, phone, specimen])

f.close()
wb.close()